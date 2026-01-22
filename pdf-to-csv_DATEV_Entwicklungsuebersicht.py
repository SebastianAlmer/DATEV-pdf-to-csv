#!/usr/bin/env python3
"""
DATEV BWA Jahresentwicklungsuebersicht (eine Seite) -> CSV.

Usage:
  python pdf-to-csv_DATEV_Entwicklungsuebersicht.py --pdf "input/BWA 2025.12.pdf" --out output/jahresentwicklung_2025_12.csv
  python pdf-to-csv_DATEV_Entwicklungsuebersicht.py --batch --input-dir input --output-dir output

- Liest genau die angegebene Seite (1-basiert) aus der PDF.
- Ohne --page wird die Seite mit "Entwicklungsuebersicht" gesucht.
- Erwartet Monats-Spalten im Kopf (auch Teiljahre wie 4 Monate + Summe) und haelt das DE-Zahlenformat bei.
- Trenner ist Semikolon; Ausgabe nutzt UTF-8 mit BOM, damit Excel Umlaute korrekt oeffnet.
- Mit --batch werden alle PDFs im Input-Ordner verarbeitet.
- Optional: Struktur-CSV kann zur Vorgabe der Zeilenstruktur genutzt werden.
- Standardmaessig wird zusaetzlich eine .xlsx geschrieben (mit --no-excel deaktivierbar).
- Ohne Parameter laeuft ein Batch von input nach output.
"""
import argparse
import csv
import difflib
import re
import sys
from pathlib import Path

import pdfplumber

MONTHS_RE = re.compile(
    r"\b(?:Jan(?:uar)?|Feb(?:ruar)?|M(?:ärz?|aerz?|rz|ar)|Apr(?:il)?|Mai|Jun(?:i)?|"
    r"Jul(?:i)?|Aug(?:ust)?|Sep(?:t(?:ember)?)?|Okt(?:ober)?|Nov(?:ember)?|Dez(?:ember)?)\.?"
    r"\s*(?:[/\.\-]\s*|\s+)\d{2,4}\b"
)
DE_NUMBER_RE = re.compile(r"-?\d{1,3}(?:\.\d{3})*,\d{2}|-?0,00")
ENTWICKLUNGSUEBERSICHT_TERMS = (
    "entwicklungsübersicht",
    "entwicklungsuebersicht",
    "jahresübersicht",
    "jahresuebersicht",
)
SEPARATOR_CLEAN_RE = re.compile(r"\s*([/.\-])\s*")

SECTION_BREAK_AFTER = {
    "Aktivierte Eigenleistungen",
    "Gesamtleistung",
    "Material-/Wareneinkauf",
    "Rohertrag",
    "So. betr. Erlöse",
    "Betrieblicher Rohertrag",
    "Gesamtkosten",
    "Betriebsergebnis",
    "Neutraler Aufwand",
    "Neutraler Ertrag",
    "Kontenklasse unbesetzt",
    "Ergebnis vor Steuern",
    "Steuern Einkommen u. Ertrag",
    "Vorläufiges Ergebnis",
}
COST_LABELS = {
    "Personalkosten",
    "Raumkosten",
    "Betriebliche Steuern",
    "Versicherungen/Beiträge",
    "Besondere Kosten",
    "Fahrzeugkosten (ohne Steuer)",
    "Werbe-/Reisekosten",
    "Kosten Warenabgabe",
    "Abschreibungen",
    "Reparatur/Instandhaltung",
    "Sonstige Kosten",
    "Gesamtkosten",
}


HEADER_FILL = "D9D9D9"
MIN_MONTH_COLUMNS = 4
STRUCTURE_MATCH_THRESHOLD = 0.86
DEFAULT_STRUCTURE_FILES = (
    "DATEV_BWA_Struktur_Vorlage.csv",
    "BWA Export Datei -leer -.csv",
    "BWA Export Datei -leer - ohne Zeile.csv",
)


def normalize_month_token(token: str):
    token = " ".join(token.split())
    token = SEPARATOR_CLEAN_RE.sub(r"\1", token)
    return token


def safe_print(message: str):
    try:
        print(message)
    except UnicodeEncodeError:
        encoding = sys.stdout.encoding or "utf-8"
        safe_message = message.encode(encoding, errors="backslashreplace").decode(
            encoding, errors="ignore"
        )
        print(safe_message)


def extract_month_tokens(text: str):
    months = [m.group(0) for m in MONTHS_RE.finditer(text)]
    seen = set()
    ordered = []
    for m in months:
        normalized = normalize_month_token(m)
        if normalized not in seen:
            seen.add(normalized)
            ordered.append(normalized)
    return ordered


def find_month_range(text: str):
    compact = " ".join(text.splitlines())
    matches = list(MONTHS_RE.finditer(compact))
    for idx in range(len(matches) - 1):
        between = compact[matches[idx].end() : matches[idx + 1].start()]
        if "-" in between:
            start = normalize_month_token(matches[idx].group(0))
            end = normalize_month_token(matches[idx + 1].group(0))
            return start, end
    return None


def build_header_columns(text: str):
    ordered = extract_month_tokens(text)
    if not ordered:
        return None
    range_tokens = find_month_range(text)
    if range_tokens:
        range_label = f"{range_tokens[0]} - {range_tokens[1]}"
        if range_label not in ordered:
            ordered.append(range_label)
    return ordered


def detect_month_header(text: str):
    lines = text.splitlines()
    for idx, raw_line in enumerate(lines):
        if "Bezeichnung" not in raw_line:
            continue
        header_text = raw_line
        if raw_line.rstrip().endswith("-") and idx + 1 < len(lines):
            header_text = f"{raw_line} {lines[idx + 1].strip()}"
        ordered = build_header_columns(header_text)
        if ordered and len(ordered) >= MIN_MONTH_COLUMNS:
            return ordered

    ordered = build_header_columns(text)
    if ordered and len(ordered) >= MIN_MONTH_COLUMNS:
        return ordered
    return None


def parse_de_amount(value):
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if DE_NUMBER_RE.fullmatch(text) is None:
        return value
    return float(text.replace(".", "").replace(",", "."))


def normalize_label(text: str) -> str:
    if not text:
        return ""
    normalized = text.strip().casefold()
    normalized = (
        normalized.replace("ä", "ae")
        .replace("ö", "oe")
        .replace("ü", "ue")
        .replace("ß", "ss")
        .replace("Ã¤", "ae")
        .replace("Ã¶", "oe")
        .replace("Ã¼", "ue")
        .replace("ÃŸ", "ss")
    )
    normalized = normalized.replace("\ufffd", "")
    normalized = re.sub(r"[^a-z0-9]+", "", normalized)
    return normalized


def parse_rows_from_text(text: str, months):
    rows = []
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        if line == "Kostenarten:":
            rows.append((line, [""] * len(months)))
            continue
        nums = DE_NUMBER_RE.findall(line)
        if len(nums) >= len(months):
            values = nums[-len(months):]
            first_num = values[0]
            pos = line.find(first_num)
            label = line[:pos].strip()
            rows.append((label, values))
    return rows


def ensure_kostenarten(rows, months):
    labels = [label for label, _ in rows]
    if "Kostenarten:" in labels:
        return rows
    first_cost_idx = next((i for i, (label, _) in enumerate(rows) if label in COST_LABELS), None)
    if first_cost_idx is None:
        return rows
    return rows[:first_cost_idx] + [("Kostenarten:", [""] * len(months))] + rows[first_cost_idx:]


def insert_section_breaks(rows, months_count):
    out = [("", [""] * months_count)]  # blank row after header
    for label, values in rows:
        out.append((label, values))
        if label in SECTION_BREAK_AFTER:
            out.append(("", [""] * months_count))
    out.append(("", [""] * months_count))  # trailing blank row
    return out


def compress_blank_rows(rows):
    cleaned = []
    prev_blank = False
    for label, values in rows:
        is_blank = not label and (not values or all(v == "" for v in values))
        if is_blank and prev_blank:
            continue
        cleaned.append((label, values))
        prev_blank = is_blank
    return cleaned


def pick_default_structure_path(base_dir: Path) -> Path | None:
    for filename in DEFAULT_STRUCTURE_FILES:
        candidate = base_dir / "DATEV Struktur" / filename
        if candidate.exists():
            return candidate
    return None


def build_output_table(header, final_rows, structure_template):
    if structure_template is not None and len(structure_template) != len(final_rows):
        raise RuntimeError(
            "Struktur hat "
            f"{len(structure_template)} Zeilen, Ergebnis hat {len(final_rows)} Zeilen."
        )

    columns = ["Bezeichnung"] + header
    rows = []
    for label, values in final_rows:
        if not label and values and all(v == "" for v in values):
            rows.append([""] + [""] * len(header))
        else:
            rows.append([label] + values)
    return columns, rows


def align_rows_to_structure(structure_template, extracted_rows, months_count):
    extracted_map = {}
    for label, values in extracted_rows:
        key = normalize_label(label)
        if not key:
            continue
        if key not in extracted_map:
            extracted_map[key] = (label, values)

    used_keys = set()
    aligned = []
    for template_label in structure_template:
        label_text = template_label or ""
        if not label_text.strip():
            aligned.append(("", [""] * months_count))
            continue
        template_key = normalize_label(label_text)
        values = None
        matched_label = None
        if template_key in extracted_map and template_key not in used_keys:
            matched_label, values = extracted_map[template_key]
            used_keys.add(template_key)
        else:
            best_key = None
            best_ratio = 0.0
            for key in extracted_map:
                if key in used_keys:
                    continue
                ratio = difflib.SequenceMatcher(None, template_key, key).ratio()
                if ratio > best_ratio:
                    best_ratio = ratio
                    best_key = key
            if best_key is not None and best_ratio >= STRUCTURE_MATCH_THRESHOLD:
                matched_label, values = extracted_map[best_key]
                used_keys.add(best_key)

        if values is None:
            values = [""] * months_count
        elif len(values) != months_count:
            values = (values + [""] * months_count)[:months_count]

        leading_ws = label_text[: len(label_text) - len(label_text.lstrip())]
        trailing_ws = label_text[len(label_text.rstrip()) :]
        core = matched_label.strip() if matched_label else label_text.strip()
        display_label = f"{leading_ws}{core}{trailing_ws}"
        aligned.append((display_label, values))
    return aligned


def write_csv_table(columns, rows, out_path: Path):
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(columns)
        for row in rows:
            writer.writerow(row)


def format_excel_sheet(ws, columns, numeric_columns, text_columns):
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    number_format = "#,##0.00"
    for col_idx, col_name in enumerate(columns, start=1):
        is_numeric = col_name in numeric_columns
        max_len = len(str(col_name))
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            cell = row[0]
            value = cell.value
            if is_numeric and isinstance(value, (int, float)):
                cell.number_format = number_format
                cell.alignment = Alignment(horizontal="right")
                display = f"{value:,.2f}"
            else:
                cell.alignment = Alignment(horizontal="left")
                display = "" if value is None else str(value)
            if display and len(display) > max_len:
                max_len = len(display)
        max_width = 60 if col_name in text_columns else 24
        width = min(max_len + 2, max_width)
        width = max(width, 8)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_excel_table(columns, rows, out_path: Path):
    try:
        import pandas as pd
    except ImportError as exc:
        raise RuntimeError("Pandas ist nicht installiert. Excel-Ausgabe nicht moeglich.") from exc

    try:
        import openpyxl  # noqa: F401
    except ImportError as exc:
        raise RuntimeError("openpyxl ist nicht installiert. Excel-Ausgabe nicht moeglich.") from exc

    out_path.parent.mkdir(parents=True, exist_ok=True)
    excel_rows = []
    for row in rows:
        label = row[0]
        values = [parse_de_amount(value) for value in row[1:]]
        excel_rows.append([label] + values)
    df = pd.DataFrame(excel_rows, columns=columns)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Daten")
        ws = writer.sheets["Daten"]
        format_excel_sheet(ws, columns, columns[1:], {"Bezeichnung"})


def find_entwicklungsuebersicht_page(pdf):
    for index, page in enumerate(pdf.pages, start=1):
        page_text = page.extract_text() or ""
        lowered = page_text.casefold()
        if any(term in lowered for term in ENTWICKLUNGSUEBERSICHT_TERMS):
            return index, page_text
    return None, ""


def convert_page_to_csv(
    pdf_path: Path,
    page_number: int | None,
    out_path: Path,
    structure_numbers: list[str] | None = None,
    write_excel_file: bool = False,
    excel_path: Path | None = None,
):
    with pdfplumber.open(pdf_path) as pdf:
        if page_number is None:
            page_number, page_text = find_entwicklungsuebersicht_page(pdf)
            if page_number is None:
                raise RuntimeError("Keine Seite mit 'Entwicklungsuebersicht' in der PDF gefunden.")
        else:
            if page_number < 1:
                raise ValueError("page_number muss 1-basiert sein")
            if page_number > len(pdf.pages):
                raise ValueError(f"PDF hat nur {len(pdf.pages)} Seiten, angefragt wurde {page_number}.")
            page_text = pdf.pages[page_number - 1].extract_text() or ""

    header = detect_month_header(page_text)
    if not header:
        raise RuntimeError("Konnte keine Monats-Spalten auf der Seite erkennen.")

    rows = parse_rows_from_text(page_text, header)
    rows = ensure_kostenarten(rows, header)

    seen = set()
    uniq_rows = []
    for label, values in rows:
        key = (label, tuple(values))
        if key not in seen:
            seen.add(key)
            uniq_rows.append((label, values))

    structure_template = structure_numbers
    if structure_template is None:
        default_structure = pick_default_structure_path(Path(__file__).resolve().parent)
        if default_structure is not None:
            structure_template = load_structure_template(default_structure)

    if structure_template is not None:
        final_rows = align_rows_to_structure(structure_template, uniq_rows, len(header))
    else:
        final_rows = compress_blank_rows(insert_section_breaks(uniq_rows, len(header)))

    columns, rows = build_output_table(header, final_rows, structure_template)
    write_csv_table(columns, rows, out_path)
    if write_excel_file:
        excel_target = excel_path or out_path.with_suffix(".xlsx")
        write_excel_table(columns, rows, excel_target)

    return out_path


def build_output_paths(out_dir: Path, pdf_path: Path, excel_dir: Path | None, write_excel_file: bool):
    csv_path = out_dir / f"{pdf_path.stem}_jahresentwicklung.csv"
    excel_path = None
    if write_excel_file:
        target_dir = excel_dir or out_dir
        excel_path = target_dir / f"{pdf_path.stem}_jahresentwicklung.xlsx"
    return csv_path, excel_path


def load_structure_template(structure_path: Path):
    with structure_path.open(encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f, delimiter=";")
        header = next(reader, None)
        if not header or not any(cell.strip() for cell in header):
            raise RuntimeError(f"Ungueltige Struktur-CSV: {structure_path}")
        label_index = None
        for idx, cell in enumerate(header):
            if "bezeichnung" in cell.strip().casefold():
                label_index = idx
                break
        if label_index is None:
            label_index = 1 if len(header) > 1 else 0
        labels = []
        for row in reader:
            if not row:
                labels.append("")
                continue
            label = row[label_index] if len(row) > label_index else ""
            labels.append(label)
    return labels


def convert_batch(
    input_dir: Path,
    out_dir: Path,
    page_number: int | None,
    write_excel_file: bool,
    excel_dir: Path | None,
    structure_numbers: list[str] | None = None,
):
    pdf_paths = sorted(input_dir.glob("*.pdf"))
    if not pdf_paths:
        raise RuntimeError(f"Keine PDF-Dateien in {input_dir} gefunden.")

    out_dir.mkdir(parents=True, exist_ok=True)
    if write_excel_file and excel_dir is not None:
        excel_dir.mkdir(parents=True, exist_ok=True)
    written = []
    skipped = []
    for pdf_path in pdf_paths:
        out_path, excel_path = build_output_paths(out_dir, pdf_path, excel_dir, write_excel_file)
        if out_path.exists() and (not write_excel_file or (excel_path and excel_path.exists())):
            safe_print(f"INFO: Uebersprungen (bereits vorhanden): {out_path}")
            continue
        try:
            convert_page_to_csv(
                pdf_path,
                page_number,
                out_path,
                structure_numbers,
                write_excel_file,
                excel_path,
            )
            written.append(out_path)
            safe_print(f"Geschrieben: {out_path}")
            if write_excel_file and excel_path is not None:
                safe_print(f"Geschrieben: {excel_path}")
        except Exception as exc:
            skipped.append((pdf_path, str(exc)))
            safe_print(f"Uebersprungen: {pdf_path} ({exc})")
    return written, skipped


def pick_default_pdf():
    candidates = sorted(Path("input").glob("*.pdf"))
    for p in candidates:
        if "bwa" in p.stem.lower():
            return p
    return candidates[0] if candidates else None


def main():
    parser = argparse.ArgumentParser(
        description="DATEV BWA Jahresentwicklungsuebersicht einer Seite in CSV umwandeln"
    )
    parser.add_argument("--pdf", type=Path, default=None, help="Pfad zur PDF-Datei")
    parser.add_argument(
        "--batch",
        action="store_true",
        help="Alle PDFs im Input-Ordner verarbeiten",
    )
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=Path("input"),
        help="Input-Ordner fuer --batch",
    )
    parser.add_argument(
        "--page",
        type=int,
        default=None,
        help="1-basierte Seitennummer; ohne Angabe wird nach 'Entwicklungsuebersicht' gesucht",
    )
    parser.add_argument(
        "--out",
        type=Path,
        default=Path("output") / "jahresentwicklung.csv",
        help="Pfad zur Ausgabe-CSV",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("output"),
        help="Ausgabe-Ordner fuer --batch",
    )
    parser.add_argument(
        "--excel-dir",
        type=Path,
        default=None,
        help="Excel-Ausgabeordner fuer --batch (Standard: output-dir)",
    )
    parser.add_argument(
        "--structure",
        type=Path,
        default=None,
        help="Pfad zur Struktur-CSV fuer die Zeilenanzahl-Pruefung",
    )
    parser.add_argument(
        "--excel",
        type=Path,
        default=None,
        help="Pfad zur Excel-Ausgabe (Einzelmodus)",
    )
    parser.add_argument(
        "--no-excel",
        action="store_true",
        help="Excel-Ausgabe deaktivieren",
    )
    raw_args = sys.argv[1:]
    if not raw_args:
        raw_args = ["--batch"]
    args = parser.parse_args(raw_args)

    structure_path = None
    if args.structure is not None:
        structure_path = args.structure
        if not structure_path.exists():
            parser.error(f"Struktur-CSV nicht gefunden: {structure_path}")
    else:
        default_structure = pick_default_structure_path(Path("."))
        if default_structure is not None:
            structure_path = default_structure

    structure_template = None
    if structure_path is not None:
        structure_template = load_structure_template(structure_path)

    write_excel_file = not args.no_excel

    if args.batch:
        input_dir = args.input_dir
        if args.pdf is not None:
            if args.pdf.is_dir():
                input_dir = args.pdf
            else:
                parser.error("--batch erwartet einen Ordner; --pdf zeigt auf eine Datei.")
        if not input_dir.exists():
            parser.error(f"Input-Ordner nicht gefunden: {input_dir}")
        _, skipped = convert_batch(
            input_dir,
            args.output_dir,
            args.page,
            write_excel_file,
            args.excel_dir,
            structure_template,
        )
        if skipped:
            safe_print(f"{len(skipped)} PDFs uebersprungen.")
        return

    pdf_path = args.pdf or pick_default_pdf()
    if pdf_path is None:
        parser.error("Keine PDF angegeben und keine Datei in ./input gefunden.")
    if not pdf_path.exists():
        parser.error(f"PDF nicht gefunden: {pdf_path}")

    out_path = convert_page_to_csv(
        pdf_path,
        args.page,
        args.out,
        structure_template,
        write_excel_file,
        args.excel,
    )
    safe_print(f"Geschrieben: {out_path}")
    if write_excel_file:
        excel_path = args.excel or args.out.with_suffix(".xlsx")
        safe_print(f"Geschrieben: {excel_path}")


if __name__ == "__main__":
    main()
